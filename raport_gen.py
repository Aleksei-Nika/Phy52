from abc import ABC, abstractmethod
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font

class ReportGenerator(ABC):
    def __init__(self, data, filename):
        self.data = data
        self.filename = filename
    
    @abstractmethod
    def ganerate(self):
        pass

    @abstractmethod
    def save(self):
        pass

    def run(self):
        print('Начало генерации')
        self.ganerate()
        self.save()
        print('Конец генерации')

class PDFReport(ReportGenerator):
    def __init__(self, data, filename):
        super().__init__(data, filename)
        self.pdf = FPDF()
    
    def ganerate(self):
        self.pdf.add_page()
        self.pdf.set_font('Arial', 'B', 16)
        self.pdf.cell(200, 16, 'Consalting Project Report', 0, 1 , 'C')
        self.pdf.set_font('Arial', 'B', 12)
        col_width = [20, 40, 40, 40]
        headers = list(self.data[0].keys())
        for i in range(len(headers)):
            self.pdf.cell(col_width[i], 10, headers[i], 1, 0, 'C')
        self.pdf.ln()
        self.pdf.set_font('Arial', '', 12)
        for row in self.data:
            values = list(row.values())
            for i in range(len(values)):
                self.pdf.cell(col_width[i], 10, str(values[i]), 1, 0, 'L')
            self.pdf.ln()

    def save(self):
        self.pdf.output(self.filename)
    

sample_data = [
    {"ID": 1, "Name": "Alice Smith", "Project": "Apollo", "Status": "Completed"},
    {"ID": 2, "Name": "Bob Johnson", "Project": "Borealis", "Status": "In Progress"},
    {"ID": 3, "Name": "Charlie Brown", "Project": "Cygnus", "Status": "Pending"}
]

report_pdf = PDFReport(sample_data, 'report.pdf')
report_pdf.run()